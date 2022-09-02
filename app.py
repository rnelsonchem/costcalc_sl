from io import BytesIO

import streamlit as st
import numpy as np
import pandas as pd
from openpyxl import load_workbook

import costcalc

def excel_bytes(coster):
    # Create the excel file. Can only save with the date and not the time
    # This must be done as a Bytes object, as described in the refs
    output = BytesIO()
    writer = pd.ExcelWriter(output)
    coster.excel(writer,) 
    writer.save()
    
    # Get the byte string for output and return this for saving
    proc_excel = output.getvalue()

    return proc_excel

### Main App Window and Functionality ###
# Get the reactions Excel file
st.sidebar.write('**Reaction Information**')
rxn_file = st.sidebar.file_uploader('Reaction Excel file upload', key='rxn')

if rxn_file:
    # Load the Excel file using openpyxl. We need to do this to get the sheet
    # names. Technically, this is making multiple copies of the rxn_sheet
    # data; however, these are likely to be pretty small files.
    wb = load_workbook(rxn_file)

    # Create a selection box for the desired sheet
    sheets = ['<Not Selected>'] + wb.sheetnames
    rxn_sheet = st.sidebar.selectbox('Sheet name', sheets, key='sheet')

    if rxn_sheet != '<Not Selected>':
        # Once the sheet is selected, get the values for the desired sheet.
        ws = wb[rxn_sheet].values
        # The column names will be the first line
        columns = next(ws)
        # Convert the remaining values to a DataFrame with the given column
        # headers. 
        df = pd.DataFrame(ws, columns=columns)
        # Find the unique compound names
        unique = df.Compound.unique()
        # Create a selection box to get the final product 
        unique = ['<Not Selected>'] + list(unique)
        final_product = st.sidebar.selectbox('Final product', unique, key='prod')

# Upload the materials Excel file
st.sidebar.write('**Materials Information**')
mat_file = st.sidebar.file_uploader('Route-specfic material file upload',
        key='rte_mat') 

# Once all of the data is collected, run it through the costing code
if mat_file and rxn_file \
        and (rxn_sheet != '<Not Selected>') \
        and (final_product != '<Not Selected>'):

    try:
        coster = costcalc.WebAppCost(mat_file, rxn_file, final_product, 
                                    rxn_sheet=rxn_sheet,)
    except ValueError as err:
        # Display errors here in a sensible way. ValueErrors at this stage are
        # mostly captured by the costcalc code
        err_sty = '<style type="text/css">h1 {color:red;}</style>\n'
        st.write(err_sty + '# *Input Error*', unsafe_allow_html=True)
        st.error(err)
        # Stop the app execution if an error is found.
        st.stop()

    coster.calc_cost()

    # Display a DataFrame of the results. At this time, the Streamlit
    # dataframe display can't handle empty cells, hence the fill=np.nan
    st.write('# Costing Output')
    c_str = f'The total RM cost for **{final_product}** is ' \
                f'**${coster.cost:.2f}**.'
    st.write(c_str)
    st.dataframe(coster.results(fill=np.nan))

    # Ask for a filename for Excel saving
    fname = st.text_input('Prove a file name below to download an Excel file:')
    if fname:
        # Add the file extension if not given
        if not fname.endswith('.xlsx'):
            fname += '.xlsx'

        st.download_button('Download', excel_bytes(coster), fname)

### References:
# See: https://stackoverflow.com/questions/36814050/openpyxl-get-sheet-by-name
# See: https://www.soudegesu.com/en/post/python/pandas-with-openpyxl/
# https://discuss.streamlit.io/t/download-button-for-csv-or-xlsx-file/17385/2
